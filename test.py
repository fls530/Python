# """
# 构造函数
#
# """
#
#
# class Football:
#
#     # 构造函数,创建对象是自动执行的方法
#     # 执行方法但是不会返回某个值
#     def __init__(self, name, age):
#         self.name = name
#         self.age = age
#         print('我叫:', name, '今年:', age, '岁')
#
#     # 定义成员变量
#     name = 'test'
#     age = 20
#
#     # 定义类方法必须要加上self形参
#     def speak(self):
#         print('name:', self.name)
#         print('age:', str(self.age))
#
#
# student = Football('马冬梅', 18)
# student.speak()
# #
#
# class People0:
#     def __init__(self, sex):
#         self.sex = sex  # 实例变量
#
#     def sayHello(self):
#         print('Sex:', self.sex)
#
#
# people = People0('man')
# people.sayHello()
#
#
# class People1:
#     sum = 0
#
#     def __init__(self, sex):
#         self.sex = sex
#         People1.sum += 1
#         print('现在有%d个人.' % People1.sum)
#
#     def sayHello(self):
#         print('Hello!')
#
#
# people1 = People1('man')
# people2 = People1('woman')

# class People:
#     sum = 0
#
#     def __init__(self, sex):
#         self.sxe = sex
#         self.__class__.sum += 1
#         print('现在有%d个人.' % self.__class__.sum)
#
#     def sayHello(self):
#         print('hello')
#
#     @classmethod
#     def plus(cls):
#         cls.sum += 1
#         print('现在有%d个人.' % cls.sum)
#
#
# people = People('man')
# People.plus()
# people2 = People('woman')
# People.plus()


# def runoob():
#     return
#
#
# try:
#     runoob()
# except AssertionError as error:
#     print(error)
# else:
#     try:
#         with open('file.log') as file:
#             read_date = file.read()
#     except FileNotFoundError as fnf_error:
#         print(fnf_error)
# finally:
#     print('这句话,无论异常是否都执行.')

# class MyError(Exception):
#     def __init__(self, value):
#         self.value = value
#
#     def __str__(self):
#         return repr(self.value)
#
#
# try:
#     raise MyError(2 * 2)
# except MyError as e:
#     print('Myexception occurred, value:', e.value)

# try:
#     raise KeyboardInterrupt
# finally:
#     print('Goodbye,world!')

# import os
# import shutil
#
# #定义复制函数
# print('-----第一题-----')
#
#
# def file_copy(file_path):
#     # 获取参数路径的文件列表
#     file_names = os.listdir(file_path)
#     # 获取当前路径父路径的绝对路径,作为复制的目标路径
#     res = os.path.dirname(os.path.abspath(__file__))
#     # 判断参数路径是否为空
#     if file_names is None:
#         return
#     # 循环判断参数路径列表中是文件还是文件夹
#     for i in file_names:
#         if os.path.isfile(file_path + '/' + i) is True:
#             # 是文件的话直接复制到目标路径下
#             shutil.copy(file_path + '/' + i, res)
#
#
# file_copy(r'D:\test')
#
#
# print('-----第二题-----')
#
#
# def homeWork():
#     # 读取case.txt
#     with open('case.txt', 'r', encoding='utf8') as f:
#         c = f.readlines()
#         data = []
#         for i in range(len(c)):
#             # 文本里读取到的值替换换行符为空
#             st = c[i].replace('\n', '')
#             # 替换为空后以','分割,赋值
#             data1 = st.split(',')
#             dic = {}
#             for j in range(len(data1)):
#                 # 再次以':'分割,赋值
#                 data2 = data1[j].split(':')
#                 # 将键和值添加进字典
#                 dic[data2[0]] = data2[1]
#             # 将添加好的字典循环添加到列表里
#             data.append(dic)
#         print(data)
#
#
# homeWork()


# def add():
#     total = 0
#     f = 100
#     for i in range(1, 11):
#         total += f
#         if i != 1:
#             f = f / 2
#     print(total)
#
#
# add()
#
#
# def mo():
#     total = 1
#     for i in range(1, 10):
#         total = (total+1)*2
#     print(total)
#
#
# mo()

# a, b = 0, 1
# for i in range(1, 11):
#     if i == 10:
#         print('第10个月：%s只兔子' % b)
#     a, b = b, a + b

# class Student:
#     def __init__(self, name, score):
#         self.name = name
#         self.score = score
#
#     def print_score(self):
#         print('%s:%s' % (self.name, self.score))
#
#
# p = Student(name='lili', score=99)
# p.print_score()

# 定义一个类
# class People:
#     # 定义类属性
#     name = ''
#     age = 0
#     # 定义类的私有属性
#     __weight = 0
#
#     # 定义实例方法
#     def __init__(self, name, age, weight):
#         self.name = name
#         self.age = age
#         self.__weight = weight
#
#     #
#     def speak(self):
#         print('%s说:我 %d 岁了' % (self.name, self.age))
#
#
# # 单继承示例
# class student(People):
#     grade = ''
#
#     def __init__(self, name, age, weight):
#         People.__init__(self, name, age, weight)
#         self.grade = g
#
#     def speak(self):
#         print('%s说:我%d岁了,我在读%s年级' % (self.name, self.age, g))
#
#
# s = student('ken', 18, 3)
# s.speak()

# 打印当前文件的绝对路径

# 获取项目目录路径
#

# import requests
# import urllib.parse
# import json
# import http.client
# import urllib.request
# import twilio
#
#
# def robot_api(keyword):
#     """机器人接口"""
#     msg_dict = {'msg': keyword}
#     msg_encode = str(urllib.parse.urlencode(msg_dict))
#     url = f'http://api.qingyunke.com/api.php?key=free&appid=0&{msg_encode}'
#     try:
#         r = requests.get(headers=headers, url=url)
#         if r.text:
#             result_dict = json.loads(r.text)
#             message = f"{result_dict['context']} \n -来自儿子的问候!"
#             send_sms(message)
#         else:
#             print('青云客api返回信息为空!')
#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#
#
# def send_sms(message):
#     """发送短信"""
#     client = client(account_sid, auth_token)
#     for phone in to_phones:
#         result = client.messages.create(
#             to=phone,
#             from_=from_phone,
#             body=message)
#         if result.sid:
#             print(f"短信发送成功:'{phone}'")
#         else:
#             print(f"发送短信失败,收信人手机号:{phone}")


# from twilio.rest import Client
#
# # Your Account SID from twilio.com/console
# account_sid = "AC718305df0baba8f0111f74228d74a6c8"
# # Your Auth Token from twilio.com/console
# auth_token = "e62053051372e79c888997509484ceee"
#
# client = Client(account_sid, auth_token)
#
# message = client.messages.create(
#     to="+8618562576336",
#     from_="+1 650 900 8670",
#     body="臭媳妇!俺爱你,生个胖肉肉")
#
# print(message.sid)

import time
import openpyxl

# from selenium import webdriver
#
# driver = webdriver.Firefox(executable_path=r'C:\Program Files\Mozilla Firefox\geckodriver.exe')
# driver.get("http://mail.126.com")
# time.sleep(5)
# driver.find_element_by_class_name('u-login-entry').click()

# 将指定的excel,加载为一个workbook对象
# wb = openpyxl.load_workbook("Cases.xlsx")
# # 选中工作簿中的表单对象
# sh = wb["Sheet1"]
# # 读取指定格子中的数据
# c12 = sh.cell(row=1, column=3).value
# print(c12)
#
# # 写入数据
# sh.cell(row=3, column=1, value='python27666')
# wb.save('Cases.xlsx')
# # 将工作簿对象保存为文件,写入的时候必须保存
# c13 = sh.cell(row=3, column=1).value
# print(c13)
# # 获取表单中的最大行
# max_row = sh.max_row
# print(max_row)
#
# # 获取表单中的最大列
# max_col = sh.max_column
# print(max_col)
#
# #rows 按行获取整个表单的所有的格子对象,每行的数据,放在一个元组中
#
# res1 = list(sh.rows)
# for i in res1:
#     for j in i:
#         print(j.value, end=' ')
#     print()
#
# #columns: 按列获取整个表单的所有的格子对象,每列数据,放在一个元组中
# res2 = list(sh.columns)
# print(res2)
#
# #需求1:读取第三行的数据
# res3 = sh.rows
# print(res3[2])
# for i in res3[2]:
#     print(i.value, end=',')
# print()
#
# # 需求2:读取第二列到第四列的数据
# res4 = list(sh.columns)
# print(res4[1:3])
#
# res5 = [j.value for x in res4[1:3] for j in x]
# print(res5)

# wb = openpyxl.load_workbook('HomeWork/Cases.xlsx')
# sh = wb['login']
# dt = list(sh.rows)
# title = []
# for item in dt[0]:
#     title.append(item.value)
# datas = []
# for item1 in dt[1:]:
#     values = []
#     for i in item1:
#         values.append(i.value)
#     cases = dict(zip(title, values))
#     datas.append(cases)
# print(datas)
import time

now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
now2 = time.strftime("%Y-%m-%d", time.localtime())
print (now)
print (now2)


