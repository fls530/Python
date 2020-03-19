import requests
import json
import openpyxl


# json_url = 'https://raw.githubusercontent.com/muxuezi/btc/master/btc_close_2017.json'
# response = requests.get(json_url)
# with open('btc_close_2017.json', 'w') as f:
#     f.write(response.text)
# file_urllib = response.json()
# print(file_urllib)
# cols = []
# filename = 'btc_close_2017.json'
# with open(filename) as f:
#     btc_data = json.load(f)
# for btc_dict in btc_data:
#     date = btc_dict['date']
#     month = btc_dict['month']
#     week = btc_dict['week']
#     weekday = btc_dict['weekday']
#     close = btc_dict['close']
#     cols.append(btc_dict)
# print(cols)


def write_excel_xlsx(path, sheet_name, cols):
    index = len(cols)
    li = ['date', 'month', 'week',  'weekday', 'close']
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        n = 0
        for j in li:
            sheet.cell(row=i + 1, column=n + 1, value=cols[i][j])
            n += 1
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def write_excel_xlsx_l(path, sheet_name, cols):
    index = len(cols)
    li = ['date', 'month', 'week', 'weekday', 'close']
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, len(li)):
        n = 0
        for j in li:
            sheet.cell(row=i + 1, column=n + 1, value=cols[i][j])
            n += 1
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")



cols = []
filename = 'btc_close_2017.json'
with open(filename) as f:
    btc_data = json.load(f)
    li = {'date': 'date', 'month': 'month', 'week': 'week', 'weekday': 'weekday', 'close': 'close'}
    cols.append(li)
for btc_dict in btc_data:
    date = btc_dict['date']
    month = btc_dict['month']
    week = btc_dict['week']
    weekday = btc_dict['weekday']
    close = btc_dict['close']

    cols.append(btc_dict)
print(cols)

book_name_xlsx = 'xlsx格式测试工作簿.xlsx'

sheet_name_xlsx = 'xlsx格式测试表'

write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, cols)
